from PyQt6.QtWidgets import QDialogButtonBox
from PyQt6.QtCore import Qt
import PyQt6.QtCore as QtCore
from PyQt6.QtWidgets import QLabel, QVBoxLayout, QHBoxLayout, QPushButton, QWidget, QDialog, QFormLayout, QLineEdit, QComboBox, QMessageBox, QTableWidget, QTableWidgetItem, QInputDialog, QFileDialog

from windows.base_window import BaseWindow
from controllers.schedule_controller import ScheduleController
from controllers.db_controller import DBController
from controllers.groups_controller import GroupsController
from controllers.teachers_controller import TeachersController
from controllers.reports_controller import ReportsController
import pandas as pd

class AdminWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Админ панель")

        self.schedule_controller = ScheduleController()
        self.db_controller = DBController()
        self.groups_controller = GroupsController(self.db_controller)
        self.teachers_controller = TeachersController()

        self.init_ui()

    def init_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        if not self.central_widget.layout():
            self.layout = QVBoxLayout()
            self.central_widget.setLayout(self.layout)

        self.label = QLabel("Добро пожаловать, Администратор!")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(self.label)

        self.manage_teachers_button = QPushButton("Управление преподавателями")
        self.manage_teachers_button.clicked.connect(self.open_manage_teachers_dialog)
        self.layout.addWidget(self.manage_teachers_button)

        self.manage_groups_button = QPushButton("Управление группами")
        self.manage_groups_button.clicked.connect(self.open_manage_groups_dialog)
        self.layout.addWidget(self.manage_groups_button)

        self.reports_button = QPushButton("Отчёты")
        self.reports_button.clicked.connect(self.open_reports_dialog)
        self.layout.addWidget(self.reports_button)

        self.logout_layout = QHBoxLayout()
        self.layout.addLayout(self.logout_layout)

        self.logout_layout.addStretch()

        self.logout_button = QPushButton("Выйти с аккаунта")
        self.logout_button.clicked.connect(self.handle_logout)
        self.logout_layout.addWidget(self.logout_button)

    def open_manage_teachers_dialog(self):
        dialog = ManageTeachersDialog(self, db_controller=self.db_controller)
        dialog.exec()

    def open_manage_groups_dialog(self):
        dialog = ManageGroupsDialog(self)
        dialog.exec()

    def open_reports_dialog(self):
        dialog = ReportsDialog(self, db_controller=self.db_controller)
        dialog.exec()

    def handle_logout(self):
        from windows.login_window import LoginWindow
        self.login_window = LoginWindow()
        self.login_window.show()
        self.close()


class CreateAccountDialog(QDialog):
    student_created = QtCore.pyqtSignal()

    def __init__(self, parent=None, role=None, group_id=None, db_controller=None):
        super().__init__(parent)
        self.setWindowTitle("Создать аккаунт")
        self.setGeometry(300, 300, 400, 300)

        self.role = role
        self.group_id = group_id
        self.db_controller = db_controller

        layout = QFormLayout()

        self.login_input = QLineEdit()
        self.first_name_input = QLineEdit()
        self.middle_name_input = QLineEdit()
        self.last_name_input = QLineEdit()

        layout.addRow("Логин:", self.login_input)
        layout.addRow("Имя:", self.first_name_input)
        layout.addRow("Фамилия:", self.last_name_input)
        layout.addRow("Отчество:", self.middle_name_input)

        save_button = QPushButton("Создать аккаунт")
        save_button.clicked.connect(self.create_account)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def create_account(self):
        login = self.login_input.text().strip()
        first_name = self.first_name_input.text().strip()
        middle_name = self.middle_name_input.text().strip()
        last_name = self.last_name_input.text().strip()
        group = self.group_id if self.role == "student" else None

        if len(login) < 4:
            QMessageBox.critical(self, "Ошибка", "Длина логин должен быть не менее 4 символов")
            return

        if not login or not first_name or not last_name:
            QMessageBox.critical(self, "Ошибка", "Все поля должны быть заполнены")
            return

        password = self.db_controller.create_account(login, first_name, middle_name, last_name, self.role, group)
        if password:
            QMessageBox.information(self, "Аккаунт создан", f"Аккаунт успешно создан. Код для завершения регистрации: {password}")
            self.student_created.emit()
            self.accept()
        else:
            QMessageBox.critical(self, "Ошибка", "Ошибка при создании аккаунта")

class ManageGroupsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Управление группами")
        self.setGeometry(300, 300, 400, 300)

        self.groups_controller = parent.groups_controller
        self.db_controller = parent.db_controller

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.group_table = QTableWidget()
        self.group_table.setColumnCount(2)
        self.group_table.setHorizontalHeaderLabels(["ID", "Название группы"])
        self.group_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.group_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.group_table.itemSelectionChanged.connect(self.on_group_selected)
        self.layout.addWidget(self.group_table)

        self.create_group_button = QPushButton("Создать группу")
        self.create_group_button.clicked.connect(self.create_group)
        self.layout.addWidget(self.create_group_button)

        self.create_student_button = QPushButton("Добавить студента")
        self.create_student_button.setEnabled(False)
        self.create_student_button.clicked.connect(self.open_create_student_account_dialog)
        self.layout.addWidget(self.create_student_button)

        self.view_students_button = QPushButton("Посмотреть студентов")
        self.view_students_button.setEnabled(False)
        self.view_students_button.clicked.connect(self.view_students)
        self.layout.addWidget(self.view_students_button)

        self.edit_group_button = QPushButton("Редактировать название группы")
        self.edit_group_button.setEnabled(False)
        self.edit_group_button.clicked.connect(self.edit_group_name)
        self.layout.addWidget(self.edit_group_button)

        self.manage_teachers_button = QPushButton("Преподаватели")
        self.manage_teachers_button.setEnabled(False)
        self.manage_teachers_button.clicked.connect(self.open_manage_group_teachers_dialog)
        self.layout.addWidget(self.manage_teachers_button)

        self.manage_schedule_button = QPushButton("Управление расписанием")
        self.manage_schedule_button.setEnabled(False)
        self.manage_schedule_button.clicked.connect(self.open_manage_group_schedule_dialog)
        self.layout.addWidget(self.manage_schedule_button)
        self.update_group_list()

        self.update_group_list()

    def update_group_list(self):
        self.group_table.setRowCount(0)
        groups = self.groups_controller.get_groups()
        for group in groups:
            row_position = self.group_table.rowCount()
            self.group_table.insertRow(row_position)
            id_item = QTableWidgetItem(str(group[0]))
            id_item.setFlags(id_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            name_item = QTableWidgetItem(group[1])
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.group_table.setItem(row_position, 0, id_item)
            self.group_table.setItem(row_position, 1, name_item)

    def on_group_selected(self):
        selected_items = self.group_table.selectedItems()
        self.create_student_button.setEnabled(bool(selected_items))
        self.view_students_button.setEnabled(bool(selected_items))
        self.edit_group_button.setEnabled(bool(selected_items))
        self.manage_teachers_button.setEnabled(bool(selected_items))
        self.manage_schedule_button.setEnabled(bool(selected_items))

    def create_group(self):
        group_name, ok = QInputDialog.getText(self, "Создать группу", "Название группы:")
        if ok and group_name:
            self.groups_controller.create_group(group_name)
            self.update_group_list()
            QMessageBox.information(self, "Успех", "Группа успешно создана")
        else:
            QMessageBox.critical(self, "Ошибка", "Введите название группы")

    def open_create_student_account_dialog(self):
        selected_items = self.group_table.selectedItems()
        if selected_items:
            group_id = int(selected_items[0].text())
            dialog = CreateAccountDialog(self, role="student", group_id=group_id, db_controller=self.db_controller)
            dialog.student_created.connect(self.update_student_list)
            dialog.exec()
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите группу")

    def view_students(self):
        selected_items = self.group_table.selectedItems()
        if selected_items:
            group_id = int(selected_items[0].text())
            dialog = ViewStudentsDialog(self, group_id=group_id)
            dialog.exec()
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите группу")

    def edit_group_name(self):
        selected_items = self.group_table.selectedItems()
        if selected_items:
            group_id = int(selected_items[0].text())
            current_name = selected_items[1].text()
            new_name, ok = QInputDialog.getText(self, "Редактировать название группы", "Новое название группы:", text=current_name)
            if ok and new_name:
                self.groups_controller.update_group_name(group_id, new_name)
                self.update_group_list()
                QMessageBox.information(self, "Успех", "Название группы успешно обновлено")
            else:
                QMessageBox.critical(self, "Ошибка", "Введите новое название группы")

    def open_manage_group_teachers_dialog(self):
        selected_items = self.group_table.selectedItems()
        if selected_items:
            group_id = int(selected_items[0].text())
            dialog = ManageGroupTeachersDialog(self, group_id=group_id, db_controller=self.db_controller)
            dialog.exec()
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите группу")

    def open_manage_group_schedule_dialog(self):
        selected_items = self.group_table.selectedItems()
        if selected_items:
            group_id = int(selected_items[0].text())
            dialog = ManageGroupScheduleDialog(self, group_id=group_id, db_controller=self.db_controller)
            dialog.exec()
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите группу")

    def update_student_list(self):
        self.view_students()


class ViewStudentsDialog(QDialog):
    def __init__(self, parent=None, group_id=None):
        super().__init__(parent)
        self.setWindowTitle("Студенты группы")
        self.setGeometry(300, 300, 400, 300)

        self.group_id = group_id
        self.groups_controller = parent.groups_controller

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.students_table = QTableWidget()
        self.students_table.setColumnCount(3)
        self.students_table.setHorizontalHeaderLabels(["ID", "Имя", "Фамилия"])
        self.students_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.students_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.students_table.itemSelectionChanged.connect(self.on_student_selected)
        self.layout.addWidget(self.students_table)

        self.remove_student_button = QPushButton("Удалить студента")
        self.remove_student_button.setEnabled(False)
        self.remove_student_button.clicked.connect(self.remove_student)
        self.layout.addWidget(self.remove_student_button)

        self.transfer_student_button = QPushButton("Перенести студента в другую группу")
        self.transfer_student_button.setEnabled(False)
        self.transfer_student_button.clicked.connect(self.transfer_student)
        self.layout.addWidget(self.transfer_student_button)

        self.load_students()

    def load_students(self):
        students = self.groups_controller.get_students_in_group(self.group_id)
        self.students_table.setRowCount(len(students))
        for row, student in enumerate(students):
            id_item = QTableWidgetItem(str(student[0]))
            id_item.setFlags(id_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            name_item = QTableWidgetItem(student[1])
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            surname_item = QTableWidgetItem(student[2])
            surname_item.setFlags(surname_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.students_table.setItem(row, 0, id_item)
            self.students_table.setItem(row, 1, name_item)
            self.students_table.setItem(row, 2, surname_item)

    def on_student_selected(self):
        selected_items = self.students_table.selectedItems()
        self.remove_student_button.setEnabled(bool(selected_items))
        self.transfer_student_button.setEnabled(bool(selected_items))

    def remove_student(self):
        selected_items = self.students_table.selectedItems()
        if selected_items:
            student_id = int(selected_items[0].text())
            status = self.groups_controller.check_status(student_id)
            if status[0][0] == 0:
                print(1)
                self.groups_controller.full_remove_student_from_group(student_id)
            else:
                print(2)
                self.groups_controller.remove_student_from_group(student_id)
            self.load_students()
            QMessageBox.information(self, "Успех", "Студент успешно удалён из группы")
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите студента для удаления")

    def transfer_student(self):
        selected_items = self.students_table.selectedItems()
        if selected_items:
            student_id = int(selected_items[0].text())
            groups = self.groups_controller.get_groups()
            group_names = [group[1] for group in groups]

            dialog = QDialog(self)
            dialog.setWindowTitle("Перенос студента")
            layout = QVBoxLayout(dialog)

            combo_box = QComboBox(dialog)
            combo_box.addItems(group_names)
            layout.addWidget(combo_box)

            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                new_group_name = combo_box.currentText()
                new_group_id = next(group[0] for group in groups if group[1] == new_group_name)
                self.groups_controller.transfer_student_to_group(student_id, new_group_id)
                self.load_students()
                QMessageBox.information(self, "Успех", "Студент успешно перенесён в другую группу")
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите студента для переноса")

class ManageTeachersDialog(QDialog):
    def __init__(self, parent=None, db_controller=None):
        super().__init__(parent)
        self.setWindowTitle("Управление преподавателями")
        self.setGeometry(300, 300, 600, 400)

        self.db_controller = db_controller
        self.teachers_controller = parent.teachers_controller

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.teachers_table = QTableWidget()
        self.teachers_table.setColumnCount(3)
        self.teachers_table.setHorizontalHeaderLabels(["ID", "Имя", "Фамилия"])
        self.teachers_table.itemSelectionChanged.connect(self.on_teacher_selected)
        self.load_teachers()
        self.layout.addWidget(self.teachers_table)

        self.create_teacher_button = QPushButton("Добавить преподавателя")
        self.create_teacher_button.clicked.connect(self.open_create_teacher_account_dialog)
        self.layout.addWidget(self.create_teacher_button)

        self.manage_specializations_button = QPushButton("Специализации преподавателя")
        self.manage_specializations_button.setEnabled(False)
        self.manage_specializations_button.clicked.connect(self.open_manage_specializations_dialog)
        self.layout.addWidget(self.manage_specializations_button)

    def load_teachers(self):
        teachers = self.db_controller.get_teachers()
        self.teachers_table.setRowCount(len(teachers))

        for row, teacher in enumerate(teachers):
            id_item = QTableWidgetItem(str(teacher[0]))
            name_item = QTableWidgetItem(teacher[1])
            surname_item = QTableWidgetItem(teacher[2])

            self.teachers_table.setItem(row, 0, id_item)
            self.teachers_table.setItem(row, 1, name_item)
            self.teachers_table.setItem(row, 2, surname_item)

    def on_teacher_selected(self):
        selected_items = self.teachers_table.selectedItems()
        self.manage_specializations_button.setEnabled(bool(selected_items))

    def open_create_teacher_account_dialog(self):
        dialog = CreateAccountDialog(self, role="teacher", db_controller=self.db_controller)
        dialog.exec()
        self.load_teachers()

    def open_manage_specializations_dialog(self):
        selected_items = self.teachers_table.selectedItems()
        if selected_items:
            row = self.teachers_table.currentRow()
            teacher_id = int(self.teachers_table.item(row, 0).text())
            dialog = ManageTeacherSpecializationsDialog(self, teacher_id=teacher_id)
            dialog.exec()

class ManageTeacherSpecializationsDialog(QDialog):
    def __init__(self, parent=None, teacher_id=None):
        super().__init__(parent)
        self.setWindowTitle("Специализации преподавателя")
        self.setGeometry(300, 300, 400, 300)

        self.teacher_id = teacher_id
        self.teachers_controller = parent.teachers_controller

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.specializations_table = QTableWidget()
        self.specializations_table.setColumnCount(2)
        self.specializations_table.setHorizontalHeaderLabels(["ID", "Название специализации"])
        self.load_specializations()
        self.layout.addWidget(self.specializations_table)

        self.add_specialization_button = QPushButton("Добавить специализацию")
        self.add_specialization_button.clicked.connect(self.open_add_specialization_dialog)
        self.layout.addWidget(self.add_specialization_button)

        self.remove_specialization_button = QPushButton("Удалить специализацию")
        self.remove_specialization_button.clicked.connect(self.remove_specialization)
        self.layout.addWidget(self.remove_specialization_button)

    def load_specializations(self):
        specializations = self.teachers_controller.get_teacher_specializations(self.teacher_id)
        self.specializations_table.setRowCount(len(specializations))
        for row, specialization in enumerate(specializations):
            id_item = QTableWidgetItem(str(specialization[0]))
            id_item.setFlags(id_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            name_item = QTableWidgetItem(specialization[1])
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.specializations_table.setItem(row, 0, id_item)
            self.specializations_table.setItem(row, 1, name_item)

    def open_add_specialization_dialog(self):
        dialog = AddSpecializationDialog(self, teacher_id=self.teacher_id)
        dialog.exec()
        self.load_specializations()

    def remove_specialization(self):
        selected_items = self.specializations_table.selectedItems()
        if selected_items:
            row = self.specializations_table.currentRow()
            specialization_id = int(self.specializations_table.item(row, 0).text())
            self.teachers_controller.remove_teacher_specialization(self.teacher_id, specialization_id)
            self.load_specializations()

class AddSpecializationDialog(QDialog):
    def __init__(self, parent=None, teacher_id=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить специализацию")
        self.setGeometry(300, 300, 400, 200)

        self.teacher_id = teacher_id
        self.teachers_controller = parent.teachers_controller

        layout = QFormLayout()

        self.specialization_combo = QComboBox()
        self.load_specializations()
        layout.addRow("Название специализации:", self.specialization_combo)

        save_button = QPushButton("Добавить")
        save_button.clicked.connect(self.add_specialization)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def load_specializations(self):
        all_specializations = self.teachers_controller.get_specializations()
        teacher_specializations = self.teachers_controller.get_teacher_specializations(self.teacher_id)
        teacher_specialization_ids = {spec[0] for spec in teacher_specializations}

        for specialization in all_specializations:
            if specialization[0] not in teacher_specialization_ids:
                if isinstance(specialization[1], bytes):
                    decoded_specialization = specialization[1].decode('utf-8')
                else:
                    decoded_specialization = specialization[1]
                self.specialization_combo.addItem(decoded_specialization)

    def add_specialization(self):
        specialization_name = self.specialization_combo.currentText()
        if specialization_name:
            specialization_id = self.teachers_controller.get_specialization_id_by_name(specialization_name)
            if specialization_id:
                self.teachers_controller.add_teacher_specialization(self.teacher_id, specialization_id)
                QMessageBox.information(self, "Успешно", "Специализация успешно добавлена")
                self.accept()
            else:
                QMessageBox.critical(self, "Ошибка", "Специализация не найдена")
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите специализацию для преподавателя")

class ManageGroupTeachersDialog(QDialog):
    def __init__(self, parent=None, group_id=None, db_controller=None):
        super().__init__(parent)
        self.setWindowTitle("Преподаватели группы")
        self.setGeometry(300, 300, 400, 300)

        self.group_id = group_id
        self.db_controller = db_controller

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.teachers_table = QTableWidget()
        self.teachers_table.setColumnCount(2)
        self.teachers_table.setHorizontalHeaderLabels(["Предмет", "Преподаватель"])
        self.teachers_table.itemSelectionChanged.connect(self.update_remove_button_state)
        self.layout.addWidget(self.teachers_table)

        self.assign_teacher_button = QPushButton("Назначить преподавателя")
        self.assign_teacher_button.clicked.connect(self.open_assign_teacher_dialog)
        self.layout.addWidget(self.assign_teacher_button)

        self.remove_teacher_button = QPushButton("Удалить преподавателя")
        self.remove_teacher_button.setEnabled(False)
        self.remove_teacher_button.clicked.connect(self.remove_teacher)
        self.layout.addWidget(self.remove_teacher_button)

        self.load_teachers()

    def load_teachers(self):
        teachers = self.db_controller.get_group_teachers(self.group_id)
        self.teachers_table.setRowCount(len(teachers))
        for row, teacher in enumerate(teachers):
            subject_item = QTableWidgetItem(teacher[0])
            subject_item.setFlags(subject_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            teacher_item = QTableWidgetItem(teacher[1])
            teacher_item.setFlags(teacher_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.teachers_table.setItem(row, 0, subject_item)
            self.teachers_table.setItem(row, 1, teacher_item)

    def open_assign_teacher_dialog(self):
        dialog = AssignTeacherDialog(self, group_id=self.group_id, db_controller=self.db_controller)
        dialog.exec()
        self.load_teachers()

    def remove_teacher(self):
        selected_items = self.teachers_table.selectedItems()
        if selected_items:
            subject_item = self.teachers_table.item(self.teachers_table.currentRow(), 0)
            if subject_item:
                subject_id = self.db_controller.get_specialization_id_by_name(subject_item.text())
                self.db_controller.remove_teacher_from_group(self.group_id, subject_id)
                QMessageBox.information(self, "Успех", "Назначение преподавателя успешно удалено")
                self.load_teachers()
            else:
                QMessageBox.critical(self, "Ошибка", "Выберите предмет для удаления назначения")
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите предмет для удаления назначения")

    def update_remove_button_state(self):
        self.remove_teacher_button.setEnabled(bool(self.teachers_table.selectedItems()))

class AssignTeacherDialog(QDialog):
    def __init__(self, parent=None, group_id=None, db_controller=None):
        super().__init__(parent)
        self.setWindowTitle("Назначить преподавателя")
        self.setGeometry(300, 300, 400, 200)

        self.group_id = group_id
        self.db_controller = db_controller

        layout = QFormLayout()

        self.subject_combo = QComboBox()
        self.load_subjects()
        layout.addRow("Предмет:", self.subject_combo)

        self.teacher_combo = QComboBox()
        self.subject_combo.currentIndexChanged.connect(self.load_teachers)
        layout.addRow("Преподаватель:", self.teacher_combo)

        save_button = QPushButton("Назначить")
        save_button.clicked.connect(self.assign_teacher)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def load_subjects(self):
        subjects = self.db_controller.get_specializations()
        for subject in subjects:
            self.subject_combo.addItem(subject[1], subject[0])

    def load_teachers(self):
        self.teacher_combo.clear()
        subject_id = self.subject_combo.currentData()
        teachers = self.db_controller.get_teachers_by_specialization(subject_id)
        for teacher in teachers:
            self.teacher_combo.addItem(f"{teacher[1]} {teacher[2]}", teacher[0])

    def assign_teacher(self):
        subject_id = self.subject_combo.currentData()
        teacher_id = self.teacher_combo.currentData()
        if subject_id and teacher_id:
            if self.db_controller.is_teacher_assigned(self.group_id, subject_id):
                QMessageBox.warning(self, "Ошибка", "Этот предмет уже назначен преподавателю для этой группы.")
                return
            self.db_controller.assign_teacher_to_group(self.group_id, subject_id, teacher_id)
            QMessageBox.information(self, "Успех", "Преподаватель успешно назначен")
            self.accept()
        else:
            QMessageBox.critical(self, "Ошибка", "Выберите предмет и преподавателя")

class ManageGroupScheduleDialog(QDialog):
    def __init__(self, parent=None, group_id=None, db_controller=None):
        super().__init__(parent)
        self.setWindowTitle("Управление расписанием группы")
        self.setGeometry(300, 300, 800, 600)
        self.group_id = group_id
        self.db_controller = db_controller
        self.main_layout = QVBoxLayout()
        self.setLayout(self.main_layout)

        self.days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
        self.schedule_layouts = []

        self.init_ui()
        self.load_and_display_schedule()

    def init_ui(self):
        days_layout = QHBoxLayout()
        self.main_layout.addLayout(days_layout)

        for day in self.days:
            day_layout = QVBoxLayout()
            day_label = QLabel(day)
            day_layout.addWidget(day_label)

            for i in range(5):
                subject_combo = QComboBox()
                self.load_subjects(subject_combo, day, i + 1)
                day_layout.addWidget(subject_combo)
                self.schedule_layouts.append((day, subject_combo))

            days_layout.addLayout(day_layout)

        save_button = QPushButton("Сохранить расписание")
        save_button.clicked.connect(self.save_schedule)
        self.main_layout.addWidget(save_button)

    def load_subjects(self, combo_box, day, period):
        specializations = self.db_controller.get_specializations()
        for specialization in specializations:
            if not self.db_controller.is_teacher_busy(self.group_id, specialization[0], day, period):
                combo_box.addItem(specialization[1], specialization[0])

    def save_schedule(self):
        schedule_data = {day: [] for day in self.days}
        for day, subject_combo in self.schedule_layouts:
            subject_id = subject_combo.currentData() or None
            schedule_data[day].append(subject_id)

        self.db_controller.save_group_schedule(self.group_id, schedule_data)
        QMessageBox.information(self, "Успех", "Расписание успешно сохранено")
        self.accept()

    def load_and_display_schedule(self):
        schedule_data = self.load_schedule(self.group_id)
        self.display_schedule(schedule_data)

    def load_schedule(self, group_id):
        return self.db_controller.load_group_schedule(group_id)

    def display_schedule(self, schedule_data):
        for day, combo_box in self.schedule_layouts:
            current_day_schedule = schedule_data.get(day, [None] * 5)
            subject_id = current_day_schedule.pop(0)
            if subject_id is not None:
                combo_box.setCurrentIndex(combo_box.findData(subject_id))
            else:
                combo_box.setCurrentIndex(-1)


class ReportsDialog(QDialog):
    def __init__(self, parent=None, db_controller=None):
        super().__init__(parent)
        self.setWindowTitle("Отчёты")
        self.setGeometry(300, 300, 400, 300)
        self.reports_controller = ReportsController(db_controller.connection)
        self.days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.group_schedule_button = QPushButton("Расписание для групп")
        self.group_schedule_button.clicked.connect(self.export_group_schedule)
        layout.addWidget(self.group_schedule_button)

        self.teacher_schedule_button = QPushButton("Расписание для преподавателей")
        self.teacher_schedule_button.clicked.connect(self.export_teacher_schedule)
        layout.addWidget(self.teacher_schedule_button)

        self.inactive_users_button = QPushButton("Неактивированные пользователи")
        self.inactive_users_button.clicked.connect(self.export_inactive_users)
        layout.addWidget(self.inactive_users_button)

    def export_group_schedule(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить расписание для групп", "",
                                                   "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            schedule_data = self.reports_controller.get_all_group_schedules()
            formatted_data = self.format_group_schedule(schedule_data)
            df = pd.DataFrame(formatted_data)
            df.to_excel(file_path, index=False)

            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            for col in ws.columns:
                max_length = 10
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Расписание для групп экспортировано в {file_path}")

    def format_group_schedule(self, schedule_data):
        formatted_data = []
        current_group = None
        dataOut = []

        for row in schedule_data:
            group_name, day, *subjects = row
            if group_name != current_group:
                if current_group is not None:
                    for j in range(len(dataOut)):
                        formatted_data.append(dataOut[j])
                    formatted_data.append([''] * 6)

                dataOut = [[''] * 6 for _ in range(5)]
                formatted_data.append([group_name, 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница'])
                current_group = group_name

            for i, subject in enumerate(subjects):
                if subject:
                    dataOut[i][0] = i + 1
                    dataOut[i][1 + self.days.index(day)] = subject

        if dataOut:
            for j in range(len(dataOut)):
                if dataOut[j][0] == '':
                    dataOut[j][0] = j + 1
                formatted_data.append(dataOut[j])

        return formatted_data

    def export_teacher_schedule(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить расписание для преподавателей", "",
                                                   "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            schedule_data = self.reports_controller.get_all_teacher_schedules()
            formatted_data = self.format_teacher_schedule(schedule_data)
            df = pd.DataFrame(formatted_data)
            df.to_excel(file_path, index=False)

            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            for col in ws.columns:
                max_length = 10
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Расписание для преподавателей экспортировано в {file_path}")

    def format_teacher_schedule(self, schedule_data):
        formatted_data = []
        current_teacher = None
        dataOut = [[''] * 6 for _ in range(5)]

        for row in schedule_data:
            teacher_name, day, *periods = row
            if teacher_name != current_teacher:
                if current_teacher is not None:
                    formatted_data.append([current_teacher, 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница'])
                    formatted_data.extend(dataOut)
                    formatted_data.append([''] * 6)
                    dataOut = [[''] * 6 for _ in range(5)]

                current_teacher = teacher_name

            for i, period in enumerate(periods):
                if period:
                    dataOut[i][0] = i + 1
                    dataOut[i][1 + self.days.index(day)] = period

        if current_teacher is not None:
            formatted_data.append([current_teacher, 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница'])
            formatted_data.extend(dataOut)

        return formatted_data

    def export_inactive_users(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить неактивированные аккаунты", "",
                                                   "Excel Files (*.xlsx);;All Files (*)")
        if not file_path:
            QMessageBox.warning(self, "Ошибка", "Файл не выбран")
            return

        try:
            inactive_teachers = self.reports_controller.get_inactive_teachers()
            inactive_students = self.reports_controller.get_inactive_students()

            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:

                df_teachers = pd.DataFrame(inactive_teachers, columns=['ФИО', 'Логин', 'Код активации'])
                df_teachers.to_excel(writer, sheet_name='Неактивированные преподаватели', index=False)

                df_students = pd.DataFrame(inactive_students, columns=['ФИО', 'Логин', 'Группа', 'Код активации'])
                df_students.to_excel(writer, sheet_name='Неактивированные студенты', index=False)

                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for col_num in range(len(df_teachers.columns)):
                        worksheet.set_column(col_num, col_num, 20)
                    worksheet.set_column(3, 3, 20)

            QMessageBox.information(self, "Успех", f"Неактивированные пользователи экспортированы в {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении файла: {e}")